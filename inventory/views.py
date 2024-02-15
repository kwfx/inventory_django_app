from collections import defaultdict
import json
import os
from io import BytesIO
from typing import Any
import openpyxl
from xlsxwriter.workbook import Workbook

from django.db.models.query import QuerySet
from django.shortcuts import render, get_object_or_404, redirect, HttpResponse
from django.urls import reverse
from django.views.generic import ListView, DetailView, RedirectView, CreateView, UpdateView, FormView, TemplateView
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from .models import Inventory, Product, InventoryProductLines, ProductLotLines, Zone, SystemStock
from .forms import ProductLotLinesFormSet, InventoryFormSet, SearchForm, StockImportForm, InventoryCompareForm
from django.http import HttpResponseRedirect, HttpResponseServerError
from django.db.models import Q
from django import forms
from django.contrib import messages
from django.core import serializers
from django.views.generic.detail import SingleObjectMixin
from django.apps import apps



class ProductListView(LoginRequiredMixin, ListView):
    model = Product
    context_object_name = "product_list"
    template_name = "inventory/prod_list.html"
    ordering = ['old_ref']
    login_url = "account_login"


class StockListView(LoginRequiredMixin, ListView):
    model = SystemStock
    context_object_name = "stock_list"
    template_name = "inventory/system_stock_list.html"
    ordering = ['product__internal_ref']
    login_url = "account_login"


class InventoryListView(LoginRequiredMixin, ListView):
    model = Inventory
    context_object_name = "inventory_list"
    template_name = "inventory/inventory_list.html"
    ordering = ['zone', 'num_inventory']
    login_url = "account_login"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        zone_id = self.request.GET.get("zone")
        num_inv = self.request.GET.get("num_inv")
        product_ref = self.request.GET.get("product_ref")
        searched_product_ids = None
        if product_ref:
            searched_product_ids = list(Product.objects.all().filter(Q(internal_ref__icontains=product_ref) |
                                                           Q(old_ref__icontains=product_ref)).values_list("id", flat=True))
        search_form = SearchForm(initial={'zone': zone_id, 'num_inv': num_inv, 'product_ref': product_ref})
        choices = [('', '')]
        if zone_id:
            choices.extend([(inv.num_inventory, inv.num_inventory) for inv in Inventory.objects.all().filter(Q(zone__id=zone_id))])
        search_form.fields["num_inv"].choices = choices
        context['search_form'] = search_form
        context['searched_product_ids'] = searched_product_ids
        return context

    def get_queryset(self) -> QuerySet[Any]:
        zone_id = self.request.GET.get("zone")
        num_inventory = self.request.GET.get("num_inv")
        filter = Q()
        if zone_id:
            filter = Q(zone__id=zone_id)
            if num_inventory:
                filter &= Q(num_inventory=num_inventory)
        return self.model.objects.all().filter(filter).order_by(*self.ordering)


class ProductUpdateView(UpdateView):
    model = Product
    template_name = "inventory/product_form_view.html"
    fields = "__all__"

class ProductCreateView(CreateView):
    model = Product
    template_name = "inventory/product_form_view.html"
    fields = "__all__"

    def get_success_url(self):
        return reverse('product_list_view')

class InventoryCreateView(CreateView):
    model = Inventory
    template_name = "inventory/inv_create.html"
    fields = [
        "zone", "num_inventory", "name_agent"
    ]

    def form_valid(self, form):
        messages.add_message(self.request, messages.SUCCESS,
                             "The Inventory was added.")
        return super().form_valid(form)


class InventoryUpdateView(SingleObjectMixin, FormView):
    """
    For adding books to a Inventory, or editing them.
    """

    model = Inventory
    template_name = "inventory/inv_update.html"
    fields = [
        "zone", "num_inventory", "name_agent"
    ]

    def get(self, request, *args, **kwargs):
        # The Inventory we're editing:
        self.object = self.get_object(queryset=Inventory.objects.all())
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # The Inventory we're uploading for:
        self.object = self.get_object(queryset=Inventory.objects.all())
        return super().post(request, *args, **kwargs)

    def get_form(self, form_class=None):
        """
        Use our big formset of formsets, and pass in the Publisher object.
        """
        return InventoryFormSet(
            **self.get_form_kwargs(), instance=self.object
        )

    def form_valid(self, form):
        """
        If the form is valid, redirect to the supplied URL.
        """
        form.save()
        messages.add_message(
            self.request, messages.SUCCESS, "Changes were saved.")
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse("inventory_update", args=[self.object.id])

def delete_record(request, model_name, pk):
    if request.user.is_authenticated:
        Model = apps.get_model("inventory." + model_name)
        record = Model.objects.get(id=pk)
        record.delete()
        return HttpResponse(f'{model_name.capitalize()} {pk} has been deleted successfully.', status=200)
    return HttpResponse(f'Not allowed', status=401)

def product_details(request, pk):
    try:
        prd = Product.objects.get(id=pk)
        prd.sale_uom = prd.get_sale_uom_display()
        json_data = serializers.serialize('json', [prd])
        return HttpResponse(json_data, content_type='application/json')
    except Exception as er:
        print(er)
        return HttpResponse(er, content_type='application/json')
    

def stock_import_controller(request):
    if request.method == "POST":
        try:
            xls_file = request.FILES['xls_file']
            file_extension = os.path.splitext(xls_file.name)[1]
            if not xls_file or file_extension not in (".xls", ".xlsx"):
                raise Exception("Vous devez charger des fichiers excel (xls ou xlsx)")
            book = openpyxl.load_workbook(filename=BytesIO(xls_file.read()))
            sh = book.worksheets[0]
            if sh.max_row:
                SystemStock.objects.all().delete()
            for row in sh.iter_rows(min_row=2):
                (internal_ref, old_ref, designation, lot, exp_date,
                    qty, qty_uom, supplier, sale_uom) = [c.value for c in row]
                product = Product.objects.all().filter(internal_ref=internal_ref, old_ref=old_ref).first()
                if not product:
                    product = Product(internal_ref=internal_ref, old_ref=old_ref, 
                            designation=designation, supplier=supplier, sale_uom=sale_uom)
                    product.save()
                stock = SystemStock(product=product, lot=lot, expiration_date=exp_date, quantity_uom = qty_uom, quantity=qty)
                stock.save()
            return HttpResponseRedirect(reverse("stock_list_view"))
        except Exception as ex:
            return HttpResponseServerError(str(ex))
    return render(request, "inventory/stock_import.html", {"form": StockImportForm})    


def StockComparisonView(request, inventory_id):
    if request.method == "GET":
        inventory = Inventory.objects.get(id=inventory_id)
        lines = []
        for product_line in inventory.inventory_product_lines.all():
            for lotline in product_line.product_lot_lines.all():
                stock_line = SystemStock.objects.filter(product=product_line.product, lot=lotline.lot).first()
                quantity_system = stock_line.quantity if stock_line else 0
                lines.append({
                    "internal_ref": product_line.product.internal_ref,
                    "old_ref": product_line.product.old_ref,
                    "designation": product_line.product.designation,
                    "lot": lotline.lot,
                    "expiration_date": str(lotline.expiration_date),
                    "supplier": product_line.product.supplier,
                    "quantity_uom": lotline.quantity_uom,
                    "quantity": lotline.quantity,
                    "quantity_system": quantity_system,
                    "ecart": lotline.quantity - quantity_system,
                    "stock_line_id": stock_line.id if stock_line else None,
                    "lotline_id": lotline.id
                })
        return render(request, "inventory/stock_comparison.html", {"lines": lines})
    else:
        try:
            data = json.loads(request.body)
            for lotline_id, vals in data.items():
                lotline = ProductLotLines.objects.get(id=lotline_id)
                if vals.get("stockline_id") == 'None':
                    new_stock_line = SystemStock(
                        product=lotline.inventory_product_line.product, 
                        lot=lotline.lot,
                        quantity=vals.get("qty_system"),
                        quantity_uom=lotline.quantity_uom,
                        expiration_date=lotline.expiration_date
                    )
                    new_stock_line.save()
                else:
                    stock_line = SystemStock.objects.get(id=vals.get("stockline_id"))
                    stock_line.quantity = vals.get("qty_system")
                    stock_line.save()
        except Exception as ex:
            return HttpResponseServerError(str(ex))
        return HttpResponseRedirect(reverse("stock_comparison", args=[inventory_id]))


def ProductAutocomplete(request):
    qs = Product.objects.all()
    term = request.GET.get("term")
    if term:
        qs = qs.filter(old_ref__icontains=term)
    data = [str(r) for r in qs.values_list("id", flat=True)]
    return HttpResponse(json.dumps(data), content_type='application/json')


def InventoryComparisonView(request, inventory_id):
    if request.method == "GET":
        inventory = Inventory.objects.get(id=inventory_id)
        lines = []
        for product_line in inventory.inventory_product_lines.all():
            for lotline in product_line.product_lot_lines.all():
                stock_line = SystemStock.objects.filter(product=product_line.product, lot=lotline.lot).first()
                quantity_system = stock_line.quantity if stock_line else 0
                lines.append({
                    "internal_ref": product_line.product.internal_ref,
                    "old_ref": product_line.product.old_ref,
                    "designation": product_line.product.designation,
                    "lot": lotline.lot,
                    "expiration_date": str(lotline.expiration_date),
                    "supplier": product_line.product.supplier,
                    "quantity_uom": lotline.quantity_uom,
                    "quantity": lotline.quantity,
                    "quantity_system": quantity_system,
                    "ecart": lotline.quantity - quantity_system,
                    "stock_line_id": stock_line.id if stock_line else None,
                    "lotline_id": lotline.id
                })
        return render(request, "inventory/stock_comparison.html", {"lines": lines})
    else:
        try:
            data = json.loads(request.body)
            for lotline_id, vals in data.items():
                lotline = ProductLotLines.objects.get(id=lotline_id)
                if vals.get("stockline_id") == 'None':
                    new_stock_line = SystemStock(
                        product=lotline.inventory_product_line.product, 
                        lot=lotline.lot,
                        quantity=vals.get("qty_system"),
                        quantity_uom=lotline.quantity_uom,
                        expiration_date=lotline.expiration_date
                    )
                    new_stock_line.save()
                else:
                    stock_line = SystemStock.objects.get(id=vals.get("stockline_id"))
                    stock_line.quantity = vals.get("qty_system")
                    stock_line.save()
        except Exception as ex:
            return HttpResponseServerError(str(ex))
        return HttpResponseRedirect(reverse("stock_comparison", args=[inventory_id]))

class InventoryCompareView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/inventory_comparison.html"
    login_url = "account_login"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory_1 = self.request.GET.get("inventory_1")
        inventory_2 = self.request.GET.get("inventory_2")
        zone = self.request.GET.get("zone")
        CompareForm = InventoryCompareForm(initial={'inventory_1': inventory_1, 'inventory_2': inventory_2, 'zone': zone})
        if zone:
            CompareForm.fields["inventory_1"].queryset = Inventory.objects.all().filter(zone__id=zone)
            if inventory_1:
                CompareForm.fields["inventory_2"].queryset = Inventory.objects.all().filter(Q(zone__id=zone) & ~Q(id=inventory_1))
        elif inventory_1 and inventory_2:
            CompareForm.fields["inventory_1"].queryset = Inventory.objects.all().filter(id=inventory_1)
            CompareForm.fields["inventory_2"].queryset = Inventory.objects.all().filter(id=inventory_2)
        context['CompareForm'] = CompareForm
        context['data'] = self._get_comparison_data(inventory_1, inventory_2)
        return context

    def _get_comparison_data(self, inventory_1, inventory_2):
        data = defaultdict(list)
        if inventory_1 and inventory_2:
            inv1 = Inventory.objects.get(id=inventory_1)
            inv2 = Inventory.objects.get(id=inventory_2)
            inv1_product_ids = set(map(lambda l: l["product_id"], inv1.inventory_product_lines.all().values()))
            inv2_product_ids = set(map(lambda l: l["product_id"], inv2.inventory_product_lines.all().values()))
            for prodline in inv1.inventory_product_lines.all().filter(Q(product__id__in=(inv1_product_ids - inv2_product_ids))):
                for lotline in prodline.product_lot_lines.all():
                    lotline_values = lotline.__dict__
                    lotline_values["quantity_inv1"] = lotline_values["quantity"]
                    lotline_values["quantity_inv2"] = 0
                    lotline_values["ecart"] = lotline_values["quantity_inv1"] - lotline_values["quantity_inv2"]
                    lotline_values["quantity_uom"] = lotline.get_quantity_uom_display()
                    data[prodline.product].append(lotline_values)
            for product_id in (inv1_product_ids & inv2_product_ids):
                product = Product.objects.get(id=product_id)
                data[product] = []
                inv1_lots_qs = inv1.inventory_product_lines.all().filter(product=product).first().product_lot_lines.all()
                inv2_lots_qs = inv2.inventory_product_lines.all().filter(product=product).first().product_lot_lines.all()
                inv1_lots = set(inv1_lots_qs.values_list("lot", flat=True))
                inv2_lots = set(inv2_lots_qs.values_list("lot", flat=True))
                for lotline in inv1_lots_qs:
                    lotline_values = lotline.__dict__
                    lotline_values["quantity_inv1"] = lotline_values["quantity"]
                    lotline_inv2 = inv2_lots_qs.filter(lot=lotline.lot).first()
                    lotline_values["quantity_inv2"] = lotline_inv2.quantity if lotline_inv2 else 0
                    lotline_values["ecart"] = lotline_values["quantity_inv1"] - lotline_values["quantity_inv2"]
                    data[product].append(lotline_values)
                for lotline in inv2_lots_qs.filter(lot__in=inv2_lots-inv1_lots):
                    lotline_values = lotline.__dict__
                    lotline_values["quantity_inv1"] = 0
                    lotline_values["quantity_inv2"] = lotline.quantity
                    lotline_values["ecart"] = lotline_values["quantity_inv1"] - lotline_values["quantity_inv2"]
                    data[product].append(lotline_values)
                
            for prodline in inv2.inventory_product_lines.all().filter(Q(product__id__in=(inv2_product_ids - inv1_product_ids))):
                for lotline in prodline.product_lot_lines.all():
                    lotline_values = lotline.__dict__
                    lotline_values["quantity_inv1"] = 0
                    lotline_values["quantity_inv2"] = lotline_values["quantity"]
                    lotline_values["ecart"] = lotline_values["quantity_inv1"] - lotline_values["quantity_inv2"]
                    lotline_values["quantity_uom"] = lotline.get_quantity_uom_display()
                    data[prodline.product].append(lotline_values)
        return dict(data)


def export_data(request, model_name):
    Model = apps.get_model("inventory." + model_name)
    ids = json.loads(request.body)
    records = Model.objects.filter(id__in=ids)
    output = BytesIO()
    workbook = Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet()
    date_cell_format = workbook.add_format({'num_format': 'dd/mm/yy'})
    if model_name == 'inventory':
        cols_header = [
            (10, "Zone"),
            (20, "N° Comptage"),
            (20, "Réf Interne"),
            (20, "Ancien réf"),
            (20, "Désignation"),
            (20, "UV"),
            (20, "Fournisseur"),
            (20, "Lot"),
            (20, "Quantité"),
            (20, "Unité"),
            (20, "Date péremption") 
        ]
        for col_index, (width, value) in enumerate(cols_header):
            worksheet.set_column(col_index, col_index, width)
            worksheet.write(0, col_index, value)
        worksheet.freeze_panes(1, 0)
        row = 1
        worksheet.set_column(10, 10, cell_format=date_cell_format)
        for record in records:
            for prodline in record.inventory_product_lines.all():
                for lotline in prodline.product_lot_lines.all():
                    worksheet.write_row(row, 0, [
                        record.zone.name, record.num_inventory, prodline.product.internal_ref, 
                        prodline.product.old_ref, prodline.product.designation, prodline.product.get_sale_uom_display(), 
                        prodline.product.supplier, lotline.lot, lotline.quantity, lotline.get_quantity_uom_display(), 
                        lotline.expiration_date,    
                    ])
                    row += 1
        workbook.close()
    output.seek(0)
    return HttpResponse(output, 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )